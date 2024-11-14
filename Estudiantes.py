import openpyxl
import random
import string
from tqdm import tqdm

def cargar_excel():
    try:
        wb = openpyxl.load_workbook("estudiantes.xlsx")
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Código", "Nombre", "Apellido", "Carrera"])
        wb.save("estudiantes.xlsx")
    return wb, sheet

def generar_codigo():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))

def registrar_estudiante():
    print("Ingresar los datos del estudiante:")
    nombre = input("Nombre: ")
    apellido = input("Apellido: ")
    carrera = input("Carrera: ")
    codigo = generar_codigo()
    
    wb, sheet = cargar_excel()
    
    sheet.append([codigo, nombre, apellido, carrera])
    
    with tqdm(total=1, desc="Guardando datos", ncols=100) as pbar:
        wb.save("estudiantes.xlsx")
        pbar.update(1)
    
    print(f"Estudiante registrado exitosamente con el código: {codigo}")

def ingresar_estudiante():
    codigo = input("Ingrese su código de estudiante: ")
    
    wb, sheet = cargar_excel()
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == codigo:
            print(f"Estudiante encontrado: {row[1]} {row[2]}, Carrera: {row[3]}")
            return
    
    print("Estudiante no encontrado.")

def menu():
    while True:
        print("\nMenu:")
        print("1. Registrar estudiante")
        print("2. Ingresar (verificar registro)")
        print("3. Salir")
        
        opcion = input("Seleccione una opción: ")
        
        if opcion == "1":
            registrar_estudiante()
        elif opcion == "2":
            ingresar_estudiante()
        elif opcion == "3":
            print("Saliendo del programa...")
            break
        else:
            print("Opción no válida. Intente nuevamente.")

if __name__ == "__main__":
    menu()
